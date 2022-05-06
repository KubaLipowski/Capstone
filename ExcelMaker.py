import pandas as pd
import openpyxl as ox
from openpyxl.utils.dataframe import dataframe_to_rows

days=['NA']
stains=['NA']
r1=['NA']
r2=['NA']
r3=['NA']
r4=['NA']
nuclei=['NA']
lumen=['NA']
vessel=['NA']
times=['NA']

data = {'Day': days,
        'Stain': stains,
        'R1': r1,
        'R2': r2,
        'R3': r3,
        'R4': r4,
        'Nuclei Count': nuclei,
        'Lumen Area': lumen,
        'Vessel ID': vessel,
        'Timestamp': times
        }
df = pd.DataFrame(data)


def update_table(day,stain,ring1, ring2,ring3,ring4,nuclei_count,lumen_area,vesselid,time):
    days.append(str(day))
    stains.append(str(stain))
    r1.append(str(ring1))
    r2.append(str(ring2))
    r3.append(str(ring3))
    r4.append(str(ring4))
    nuclei.append(str(nuclei_count))
    lumen.append(str(lumen_area))
    vessel.append(str(vesselid))
    times.append(str(time))

    index = len(days)-1
    df.at[index,"Day"]=days[index]
    df.at[index,"Stain"]=stains[index]
    df.at[index,"R1"]=r1[index]
    df.at[index,"R2"]=r2[index]
    df.at[index,"R3"]=r3[index]
    df.at[index,"R4"]=r4[index]
    df.at[index,"Nuclei Count"]=nuclei[index]
    df.at[index,"Lumen Area"]=lumen[index]
    df.at[index,"Vessel ID"]=vessel[index]
    df.at[index,"Timestamp"]=times[index]
    df.sort_values(by=['Day'])
    return df


def write_to_excel(df,filename):
    wb = ox.load_workbook(filename)
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=False):
        col = 1
        ro = ws.max_row+1
        for val in r:
            if (val == 'NA'):
                pass
            else:
                ws.cell(column=col, row=ro, value=val)
                wb.save(filename)
            col = col+1


def main():
    da=update_table(1,'ED2',1,2,3,4,5,6,'1:12:13',1)
    da=update_table(9,'DAPI',1,2,3,4,5,6,'1:13:14',2)
    write_to_excel(da,'C:/Users/student/Desktop/Validation_Project/Output.xlsx')
    print(da)


if __name__ == "__main__":main()
